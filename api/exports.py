"""
Utility functions for generating PDF and Excel exports
"""
from io import BytesIO
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.utils.translation import gettext as _
from decimal import Decimal
from datetime import datetime


def format_currency(value):
    """Format decimal value as currency"""
    if isinstance(value, Decimal):
        return f"{value:,.2f}"
    return f"{float(value):,.2f}"


# ============================================================================
# PDF Export Functions
# ============================================================================

def generate_invoice_pdf(invoice):
    """Generate PDF for an invoice"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.5*inch, leftMargin=0.5*inch,
                          topMargin=0.75*inch, bottomMargin=0.75*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=6,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(_("INVOICE"), title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Invoice info
    info_data = [
        [_("Invoice Number:"), str(invoice.invoice_number), _("Date:"), invoice.invoice_date.strftime("%d/%m/%Y")],
        [_("Due Date:"), invoice.due_date.strftime("%d/%m/%Y"), _("Status:"), invoice.get_status_display()],
    ]
    info_table = Table(info_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
    info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#f0f0f0')),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Client info
    elements.append(Paragraph(_("<b>Client Information</b>"), styles['Heading3']))
    client_info = f"""
    <b>{invoice.client.name}</b><br/>
    {invoice.client.company}<br/>
    {invoice.client.address}<br/>
    {invoice.client.postal_code} {invoice.client.city}, {invoice.client.country}<br/>
    {_("Tax ID")}: {invoice.client.tax_id}<br/>
    {_("Email")}: {invoice.client.email}<br/>
    {_("Phone")}: {invoice.client.phone}
    """
    elements.append(Paragraph(client_info, styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Invoice items table
    items_data = [[_("Description"), _("Qty"), _("Unit Price"), _("Tax %"), _("Total")]]
    for item in invoice.items.all():
        items_data.append([
            item.description,
            f"{item.quantity:.2f}",
            f"${format_currency(item.unit_price)}",
            f"{item.tax_rate:.1f}%",
            f"${format_currency(item.total)}"
        ])
    
    items_table = Table(items_data, colWidths=[2.5*inch, 0.75*inch, 1.2*inch, 0.75*inch, 1.2*inch])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Totals
    totals_data = [
        [_("Subtotal"), f"${format_currency(invoice.subtotal)}"],
        [_("Tax Amount"), f"${format_currency(invoice.tax_amount)}"],
        [_("Total"), f"${format_currency(invoice.total)}"],
        [_("Paid Amount"), f"${format_currency(invoice.paid_amount)}"],
        [_("Balance Due"), f"${format_currency(invoice.total - invoice.paid_amount)}"],
    ]
    totals_table = Table(totals_data, colWidths=[4*inch, 1.5*inch])
    totals_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -2), (-1, -1), colors.HexColor('#DCE6F1')),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(totals_table)
    
    if invoice.notes:
        elements.append(Spacer(1, 0.3*inch))
        elements.append(Paragraph(_("<b>Notes</b>"), styles['Heading3']))
        elements.append(Paragraph(invoice.notes, styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_proforma_pdf(proforma):
    """Generate PDF for a proforma invoice"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.5*inch, leftMargin=0.5*inch,
                          topMargin=0.75*inch, bottomMargin=0.75*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=6,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(_("PROFORMA INVOICE"), title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Proforma info
    info_data = [
        [_("Proforma Number:"), str(proforma.proforma_number), _("Date:"), proforma.issue_date.strftime("%d/%m/%Y")],
        [_("Expiry Date:"), proforma.expiry_date.strftime("%d/%m/%Y"), _("Status:"), proforma.get_status_display()],
    ]
    info_table = Table(info_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
    info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#f0f0f0')),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Client info
    elements.append(Paragraph(_("<b>Client Information</b>"), styles['Heading3']))
    client_info = f"""
    <b>{proforma.client.name}</b><br/>
    {proforma.client.company}<br/>
    {proforma.client.address}<br/>
    {proforma.client.postal_code} {proforma.client.city}, {proforma.client.country}
    """
    elements.append(Paragraph(client_info, styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Proforma items table
    items_data = [[_("Description"), _("Qty"), _("Unit Price"), _("Tax %"), _("Total")]]
    for item in proforma.items.all():
        items_data.append([
            item.description,
            f"{item.quantity:.2f}",
            f"${format_currency(item.unit_price)}",
            f"{item.tax_rate:.1f}%",
            f"${format_currency(item.total)}"
        ])
    
    items_table = Table(items_data, colWidths=[2.5*inch, 0.75*inch, 1.2*inch, 0.75*inch, 1.2*inch])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Totals
    totals_data = [
        [_("Subtotal"), f"${format_currency(proforma.subtotal)}"],
        [_("Tax Amount"), f"${format_currency(proforma.tax_amount)}"],
        [_("Total"), f"${format_currency(proforma.total)}"],
    ]
    totals_table = Table(totals_data, colWidths=[4*inch, 1.5*inch])
    totals_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#DCE6F1')),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(totals_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


# ============================================================================
# Excel Export Functions
# ============================================================================

def generate_invoice_excel(invoice):
    """Generate Excel file for an invoice"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = _("INVOICE")
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Invoice info
    ws['A3'] = _("Invoice Number:")
    ws['B3'] = invoice.invoice_number
    ws['D3'] = _("Date:")
    ws['E3'] = invoice.invoice_date
    
    ws['A4'] = _("Client:")
    ws['B4'] = invoice.client.name
    ws['D4'] = _("Due Date:")
    ws['E4'] = invoice.due_date
    
    ws['A5'] = _("Status:")
    ws['B5'] = invoice.get_status_display()
    
    # Items header
    row = 7
    headers = [_("Description"), _("Qty"), _("Unit Price"), _("Tax %"), _("Total")]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Items
    row = 8
    for item in invoice.items.all():
        ws.cell(row=row, column=1).value = item.description
        ws.cell(row=row, column=2).value = item.quantity
        ws.cell(row=row, column=3).value = item.unit_price
        ws.cell(row=row, column=4).value = item.tax_rate
        ws.cell(row=row, column=5).value = item.total
        row += 1
    
    # Totals
    totals_row = row + 1
    ws[f'A{totals_row}'] = _("Subtotal:")
    ws[f'B{totals_row}'] = invoice.subtotal
    
    ws[f'A{totals_row+1}'] = _("Tax Amount:")
    ws[f'B{totals_row+1}'] = invoice.tax_amount
    
    ws[f'A{totals_row+2}'] = _("Total:")
    ws[f'B{totals_row+2}'] = invoice.total
    ws[f'B{totals_row+2}'].font = Font(bold=True, size=12)
    
    ws[f'A{totals_row+3}'] = _("Paid Amount:")
    ws[f'B{totals_row+3}'] = invoice.paid_amount
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_invoices_list_excel(invoices):
    """Generate Excel file with list of invoices"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        _("Invoice Number"),
        _("Client"),
        _("Date"),
        _("Due Date"),
        _("Amount"),
        _("Paid"),
        _("Status"),
        _("Created At")
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Data
    row = 2
    for invoice in invoices:
        ws.cell(row=row, column=1).value = invoice.invoice_number
        ws.cell(row=row, column=2).value = invoice.client.name
        ws.cell(row=row, column=3).value = invoice.invoice_date
        ws.cell(row=row, column=4).value = invoice.due_date
        ws.cell(row=row, column=5).value = invoice.total
        ws.cell(row=row, column=6).value = invoice.paid_amount
        ws.cell(row=row, column=7).value = invoice.get_status_display()
        ws.cell(row=row, column=8).value = invoice.created_at
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_proforma_excel(proforma):
    """Generate Excel file for proforma invoice"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Proforma"
    
    # Styling
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    title_font = Font(bold=True, size=16)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:E1')
    title = ws['A1']
    title.value = f"PROFORMA {proforma.proforma_number}"
    title.font = title_font
    
    # Company info
    ws['A3'] = "Date:"
    ws['B3'] = proforma.proforma_date
    ws['A4'] = "Valid until:"
    ws['B4'] = proforma.validity_date
    
    # Client info
    ws['A6'] = "Client Information:"
    ws['A7'] = "Name:"
    ws['B7'] = proforma.client.name
    ws['A8'] = "Email:"
    ws['B8'] = proforma.client.email
    ws['A9'] = "Address:"
    ws['B9'] = proforma.client.address
    
    # Items header
    row = 12
    headers = ["Product", "Quantity", "Unit Price", "Amount", "Tax"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Items
    row = 13
    total = 0
    for item in proforma.items.all():
        ws.cell(row=row, column=1).value = item.product.name
        ws.cell(row=row, column=2).value = item.quantity
        ws.cell(row=row, column=3).value = float(item.unit_price)
        ws.cell(row=row, column=4).value = float(item.subtotal)
        ws.cell(row=row, column=5).value = float(item.tax_amount)
        total += item.subtotal
        row += 1
    
    # Totals
    row += 1
    ws[f'C{row}'] = "Subtotal:"
    ws[f'D{row}'] = float(proforma.subtotal)
    row += 1
    ws[f'C{row}'] = "Tax:"
    ws[f'D{row}'] = float(proforma.tax_amount)
    row += 1
    ws[f'C{row}'] = "TOTAL:"
    ws[f'D{row}'] = float(proforma.total)
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
